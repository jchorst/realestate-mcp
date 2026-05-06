# ruff: noqa: E501, RUF001
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = r"C:\Users\jchorst\Desktop\OBX_Aug2026_Rentals_v5.xlsx"

TOP_PICKS = [
    (1, "Beach Bum Inn — 7 BR, Pool, Ocean Views", "Nags Head", 7418, 7, 9, 4.5, 4.81, 74, "Sat Aug 15", "Best value in the set: cheapest Tier-A listing with the deepest review history.", "https://www.airbnb.com/rooms/21766484"),
    (2, "TRULY OceanFront *Pet Friendly*", "Rodanthe", 8708, 8, 9, 4.5, 4.94, 52, "Sun Aug 16", "Highest-rated proven oceanfront under budget. Pet friendly. South-OBX (longer drive).", "https://www.airbnb.com/rooms/912371451625476383"),
    (3, "Heated Pool / Waterfront / Pier / Tennis", "Avon", 9096, 8, 11, None, 4.84, 62, "Sat Aug 15", "Long-running waterfront listing with private pier and tennis.", "https://www.airbnb.com/rooms/5567139"),
    (4, "Less than 5min Walk to Beach, Pool, Fire Pit", "Corolla", 9815, 8, 14, 7.0, 4.96, 28, "Sun Aug 16", "Highest-rated home in budget. Big bed count, short walk to beach.", "https://www.airbnb.com/rooms/867311257318543309"),
    (5, "Luxurious 7BR Oasis with Sound Views", "Corolla", 9892, 7, 11, 8.0, 4.85, 33, "Fri Aug 14", "Sound-side luxury, 8 baths, strong rating.", "https://www.airbnb.com/rooms/1038759850099623996"),
    (6, "Sea Biscuit — pool, sleeps 26, pool table", "Nags Head", 10410, 7, 14, None, 4.79, 71, "Sun Aug 16", "Massive review history, sleeps 26 — great for big crews.", "https://www.airbnb.com/rooms/34533054"),
    (7, "Windswept — hot tub, pool, oceanfront", "Nags Head", 10415, 8, 12, None, 4.84, 32, "Sun Aug 16", "Oceanfront with proven reviews. Sleeps 24.", "https://www.airbnb.com/rooms/38040310"),
    (8, "View 7BR/8BA — Elevator, Pool, Theater", "Salvo", 10445, 7, 14, 8.0, 4.94, 17, "Sun Aug 16", "Elevator (rare and helpful), home theater + game room.", "https://www.airbnb.com/rooms/572037515475344757"),
    (9, "Ferrari II — Semi-Oceanfront, Pool, Hot Tub, Arcade", "Nags Head", 7954, 8, 14, None, 4.72, 18, "Sun Aug 16", "Cheap for an 8-bedroom + arcade — strong family option.", "https://www.airbnb.com/rooms/32018897"),
    (10, "A Raye Of Sunshine — Saltwater Pool, Ocean Views", "Corolla", 7638, 7, 11, 7.5, 4.56, 9, "Fri Aug 14", "Cheapest Corolla listing with a credible rating + ocean views.", "https://www.airbnb.com/rooms/1046417072862552207"),
    (11, "Duck Location! Volleyball, Pool, Game Room", "Duck", 8612, 7, 10, 8.5, 4.83, 12, "Sun Aug 16", "Best-reviewed sub-$9k option in the Duck area.", "https://www.airbnb.com/rooms/983350702839074200"),
    (12, "Semi-Oceanfront 8 bed with Pool, Spa", "Nags Head", 10023, 8, 11, 5.5, 4.79, 29, "Sun Aug 16", "Solid all-around, semi-oceanfront with spa.", "https://www.airbnb.com/rooms/1098516403372996262"),
]

HONORABLE = [
    ("7BR Lakefront — Pool, Tennis, Dog Friendly", "Avon", 5250, 7, 9, None, None, 0, "Sun Aug 16", "Cheap and well-equipped, but no review history yet.", "https://www.airbnb.com/rooms/1539330844217519499"),
    ("Lucky Catch — 7BR Waterview, Pool, Hot Tub", "Avon", 5287, 7, 9, None, None, 0, "Sat Aug 15", "Cheap, pool/hot tub, but unreviewed.", "https://www.airbnb.com/rooms/1538719965611327641"),
    ("Salty Daze — Private Pool, Hot Tub, Game Room", "Rodanthe", 5433, 7, 10, None, None, 0, "Fri Aug 14", "Only Fri Aug 14 option in the cheap tier; unreviewed.", "https://www.airbnb.com/rooms/1561677417023280300"),
    ("Almost Heaven — Private Pool & Hot Tub, Pets OK", "Corolla", 5375, 7, 12, 4.0, 4.75, 4, "Sun Aug 16", "Cheapest in Corolla with at least some reviews (4).", "https://www.airbnb.com/rooms/794902786475184746"),
]

NOTES = [
    "Filters applied: 7+ bedrooms, 7 nights, total <= $11,000, check-in on Fri Aug 14 / Sat Aug 15 / Sun Aug 16, 2026.",
    "Geographic flavor: Corolla/Duck = developed north end. Nags Head/KDH = central, lots of restaurants.",
    "Avon/Rodanthe/Salvo = quieter southern Hatteras Island, ~1–1.5h further drive but better surf and stargazing.",
    "'—' / blank in Bathrooms means Airbnb didn't return a count in the search payload — check the listing page.",
    "Prices and availability are a snapshot as of 2026-05-05 — they will drift; verify on the listing page before booking.",
    "Excluded listings with rating < 4.0 (e.g. 'Captain Jack's Landing' at 2.5, 'Bela Vista' at 3.67) and anything > $11k total.",
    "See the 'Carolina Designs Direct', 'Twiddy Direct', and 'Surf or Sound Direct' sheets for matching listings from local OBX property managers.",
    "  - Carolina Designs and Twiddy cover north OBX (Corolla through Nags Head). Surf or Sound covers Hatteras Island (south OBX) — Rodanthe through Hatteras village.",
]


# Carolina Designs results: from scripts/fetch_carolinadesigns_aug15.py — Aug 15 2026 Saturday arrival, 7 nights.
# No public ratings/reviews exposed by their API, so ranking is by location/value/capacity, not crowd-sourced.
CD_PICKS = [
    (1, "BAREFEET RETREAT", "Nags Head", "MP 16.5", "Oceanfront", 9490, 9, 6, 1, 18, True, True, "Oceanfront 9BR for under $10k is rare. Pet friendly. Central OBX with restaurants nearby.", "https://www.carolinadesigns.com/nags-head-vacation-rental/827-barefeet-retreat/"),
    (2, "SAPPHIRE BY THE SEA", "Kill Devil Hills", "MP 7.5", "Semi-Oceanfront", 9490, 9, 9, 2, 20, True, True, "Big sleep count (20), 9 full baths, semi-oceanfront, pet friendly. Great for large group.", "https://www.carolinadesigns.com/kill-devil-hills-vacation-rental/844-sapphire-by-the-sea/"),
    (3, "OCEAN JOURNEY", "Duck", "Duck Landing", "330 yds from Beach Access", 7990, 7, 7, 2, 16, False, True, "Duck is the upscale quiet north town. 7 full + 2 half baths, private pool.", "https://www.carolinadesigns.com/duck-vacation-rental/078-ocean-journey/"),
    (4, "PELICAN COVE", "Duck", "Duck Landing", "321 yds from Beach Access", 7990, 7, 6, 1, 16, True, True, "Same price/area/sleeps as Ocean Journey but pet friendly. Pick this one if bringing dogs.", "https://www.carolinadesigns.com/duck-vacation-rental/459-pelican-cove/"),
    (5, "SANDY PAWS", "Corolla", "Sec. O, Ocean Sands", "189 yds from Beach Access", 5990, 7, 6, 1, 22, False, True, "Best $/sleep value in the set: 22 sleeps for $5,990. Short walk to beach.", "https://www.carolinadesigns.com/corolla-vacation-rental/931-sandy-paws/"),
    (6, "THE SPACE BETWEEN", "Duck", "Carolina Dunes", "225 yds from Beach Access", 9250, 7, 7, 1, 16, False, True, "Duck location, 7 full baths, closer to beach than Ocean Journey/Pelican Cove.", "https://www.carolinadesigns.com/duck-vacation-rental/414-the-space-between/"),
    (7, "THE QUARTERDECK", "Corolla", "Pine Island", "270 yds from Beach Access", 5890, 7, 5, 1, None, False, True, "Cheapest in the set. Pine Island is far-north quiet. Sleeps count not exposed by API — verify on listing.", "https://www.carolinadesigns.com/corolla-vacation-rental/450-the-quarterdeck/"),
]

CD_HEADERS = [
    "Rank", "Listing", "Town", "Subdivision", "Beach Proximity",
    "7-Night Total", "BR", "Full Baths", "Half Baths", "Sleeps",
    "Pets OK", "Private Pool", "Why", "URL",
]

# Picks: (rank, name, town, location, 7-night total, BR, full baths, half baths, listing_id, why)
SOS_PICKS = [
    (1, "1st Wave", "Waves", "Oceanfront", 6395, 7, 6, 2, 31, "Oceanfront 7BR for under $7k is rare. Smallest Hatteras-Island village = very quiet."),
    (2, "Casa Del Sol", "Rodanthe", "Semi-Oceanfront", 4395, 8, 8, 1, 676, "Semi-oceanfront 8BR + 8 full baths for $4,395. Best capacity-for-price in the set."),
    (3, "Sunset Cove", "Rodanthe", "Soundfront", 5995, 8, 8, 2, 1036, "Soundfront (sunset views over Pamlico Sound) 8BR with 10 baths for under $6k."),
    (4, "Peaceful Retreat", "Rodanthe", "Oceanview", 6495, 8, 8, 1, 534, "Ocean-view 8BR; not oceanfront but still a strong vantage. Eight full baths."),
    (5, "Pura Vida", "Avon", "Oceanfront", 9995, 8, 7, 3, 1074, "Oceanfront 8BR in Avon — the largest Hatteras-Island town with restaurants, Food Lion, kiteboarding. Most expensive but best location/town combo."),
    (6, "Flip Floppin'", "Avon", "Oceanfront", 9495, 7, 7, 0, 901, "Oceanfront Avon, slightly smaller and cheaper than Pura Vida. 7 full baths."),
    (7, "Beach Babies", "Hatteras village", "Oceanfront", 7895, 7, 5, 2, 1130, "Oceanfront in Hatteras village — but it's the most remote town, +1.5 hr from Wright bridge. Ferry to Ocracoke departs here."),
    (8, "Costa Del Mar", "Salvo", "Soundside", 6395, 9, 9, 1, 1238, "Biggest in the set: 9BR + 9 full baths for $6,395. Soundside (sound-facing, not oceanfront) — cheaper because of that."),
    (9, "Sailor's Landing", "Rodanthe", "Oceanfront", 8995, 7, 5, 1, 750, "Most expensive 7BR in the set, only 5 full baths. The cheaper 8BR Rodanthe options likely beat this on value."),
    (10, "Slainte", "Salvo", "Semi-Soundfront", 3295, 7, 7, 2, 320, "7 full + 2 half baths for $3,295 — best baths-per-dollar in the cheap tier."),
    (11, "Spittin' Kitty", "Salvo", "Soundside", 4195, 7, 7, 1, 1031, "Salvo soundside, 8 total baths under $4.2k. Quiet."),
    (12, "Carolina Daydream", "Salvo", "Soundside", 2995, 7, 5, 1, 512, "Tied for cheapest in the entire OBX search. Soundside Salvo — quiet but no ocean walkout."),
    (13, "Fire Island", "Salvo", "Semi-Soundfront", 2995, 7, 5, 1, 440, "Tied for cheapest. Same caveats as Carolina Daydream — far-south quiet, soundside."),
]

SOS_HEADERS = [
    "Rank", "Listing", "Town", "Location",
    "7-Night Total", "BR", "Full Baths", "Half Baths",
    "Why", "URL",
]

SOS_URL_BASE = "https://www.surforsound.com/hatteras-vacation-rental/property"

SOS_NOTES = [
    "Source: Surf or Sound Realty (surforsound.com) — Hatteras Island specialist (south-OBX, the area Carolina Designs and Twiddy don't cover).",
    "Coverage: Rodanthe, Waves, Salvo, Avon, Buxton, Frisco, Hatteras village. Buxton and Frisco had 0 in-budget 7+ BR for Aug 15. Ocracoke is not covered.",
    "Saturday-arrival only (Aug 15 of your three candidate dates).",
    "Methodology: searched all 7 Hatteras Island villages with 7+ BR, check_in=2026-08-15. 13 listings returned with availability — ALL came in at <= $11k. Hatteras Island weekly rates are much lower than North OBX peers.",
    "Why so much cheaper than Carolina Designs / Twiddy / Airbnb? Hatteras Island is a longer drive (extra 30-90 min south of Oregon Inlet bridge) and has less development; demand is correspondingly lower for inland-focused renters but pricing reflects that.",
    "Towns by character: Avon = largest Hatteras town with amenities. Rodanthe/Waves/Salvo = the 'tri-villages', small + quiet. Hatteras village = remote, ferry to Ocracoke. Buxton = surfer hub, lighthouse.",
    "Storm risk: NC-12 (the only road to Hatteras Island) closes during storms — travel insurance with a 'named-storm' clause is more relevant here than for North OBX.",
    "Prices are a snapshot from the search endpoint as of 2026-05-05 — they will drift; verify on the listing page before booking.",
    "Note: URLs use listing_id from the search; numeric IDs may not match what you see in the URL bar after Surf or Sound's site redirects to the slug-based form.",
]


TWIDDY_PICKS = [
    (1, "Blue Marlin I", "Duck", "Ocean Crest", "More than 500 ft from beach", 5700, 7, 7, 1, 18, False, False, False, "Only Twiddy listing in budget. Cheap for Duck — but no private pool, not oceanfront. Best $/sleep value if back-row + community pool is acceptable.", "https://www.twiddy.com/outer-banks/duck/ocean-crest/rentals/blue-marlin-i/"),
]

TWIDDY_HEADERS = [
    "Rank", "Listing", "Town", "Neighborhood", "Beach Proximity",
    "7-Night Total", "BR", "Full Baths", "Half Baths", "Sleeps",
    "Oceanfront", "Pets OK", "Private Pool", "Why", "URL",
]

TWIDDY_NOTES = [
    "Source: Twiddy & Company (twiddy.com) — one of the largest OBX property managers, north-OBX heavy, premium-leaning inventory.",
    "Coverage gaps:",
    "  - North OBX only: Corolla, Duck, Southern Shores, Kill Devil Hills, Nags Head, plus the 4x4 area (remote 4WD beach north of Corolla).",
    "  - Saturday-arrival only (same as Carolina Designs): only Sat Aug 15 of your three candidate dates.",
    "  - No public ratings exposed by their API.",
    "Methodology: searched all 368 listings with 7+ BR across the 6 north-OBX areas Twiddy covers; 14 had Aug 15 2026 availability; only 1 came in at <= $11,000 total.",
    "Why so few in-budget? Twiddy's 7+ BR portfolio skews to large oceanfront homes (10-30 BR) with peak-week prices of $15k-$50k+. Their 7-bedroom inventory tends to also be premium.",
    "Price note: Twiddy weekly_rate has two numbers per row — the discounted price you pay (first $) and the strikethrough list price. We use the discounted price.",
    "Prices are a snapshot as of 2026-05-05 — they will drift; verify on the listing page before booking.",
]


CD_NOTES = [
    "Source: Carolina Designs Realty (carolinadesigns.com) — local OBX property manager, ~350 listings. Booked direct, not via Airbnb.",
    "Important coverage gaps:",
    "  - North OBX only: Corolla, Duck, Southern Shores, Kitty Hawk, Kill Devil Hills, Nags Head. No Avon/Rodanthe/Hatteras Island.",
    "  - Saturday-arrival only: of your three candidate dates (Fri Aug 14 / Sat Aug 15 / Sun Aug 16), only Sat Aug 15 is supported.",
    "  - No ratings or reviews exposed by their API — ranking is by location/value/capacity, not crowd-sourced. Verify reputation independently.",
    "Methodology: searched all 230 listings with 7+ BR across 6 north-OBX towns; 20 had Aug 15 2026 availability; 7 came in at <= $11,000 total.",
    "All listings have private pools. 'Beach Proximity' is from Carolina Designs's own labeling (Oceanfront / Semi-Oceanfront / X yds from access).",
    "Prices are a snapshot as of 2026-05-05 — they will drift; verify on the listing page before booking.",
]

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
SECTION_FONT = Font(bold=True, size=12)
LINK_FONT = Font(color="0563C1", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

HEADERS = [
    "Rank", "Listing", "Town", "7-Night Total",
    "BR", "Beds", "Baths", "Rating", "Reviews",
    "Check-in", "Why", "URL",
]


def write_header(ws, row: int, label: str) -> int:
    ws.cell(row=row, column=1, value=label).font = SECTION_FONT
    ws.cell(row=row, column=1).fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    return row + 1


def write_column_headers(ws, row: int) -> int:
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    return row + 1


def write_row(ws, row: int, rank, name, town, total, br, beds, baths, rating, reviews, checkin, why, url):
    cells = [
        (1, rank, CENTER, None),
        (2, name, WRAP, None),
        (3, town, None, None),
        (4, total, RIGHT, "$#,##0"),
        (5, br, CENTER, None),
        (6, beds, CENTER, None),
        (7, baths, CENTER, "0.0"),
        (8, rating, CENTER, "0.00"),
        (9, reviews, CENTER, None),
        (10, checkin, CENTER, None),
        (11, why, WRAP, None),
        (12, url, None, None),
    ]
    for col, value, align, fmt in cells:
        c = ws.cell(row=row, column=col, value=value)
        c.border = BORDER
        if align is not None:
            c.alignment = align
        if fmt is not None and value is not None:
            c.number_format = fmt
    url_cell = ws.cell(row=row, column=12)
    url_cell.hyperlink = url
    url_cell.font = LINK_FONT
    url_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "OBX 7BR Aug 2026"

    row = 1
    title = ws.cell(row=row, column=1, value="Outer Banks NC — 7+ BR Rentals, 3rd Week of August 2026 (Aug 15–22)")
    title.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    row += 1

    subtitle = ws.cell(row=row, column=1, value="Sorted from most to least recommended. Filtered to ≤ $11,000 total for 7 nights.")
    subtitle.font = Font(italic=True, color="555555")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    row += 2

    row = write_header(ws, row, "Top Picks")
    row = write_column_headers(ws, row)
    for entry in TOP_PICKS:
        write_row(ws, row, *entry)
        row += 1
    row += 1

    row = write_header(ws, row, "Honorable Mentions (cheaper but little/no review history)")
    row = write_column_headers(ws, row)
    for entry in HONORABLE:
        write_row(ws, row, "", *entry)
        row += 1
    row += 2

    row = write_header(ws, row, "Notes")
    for n in NOTES:
        c = ws.cell(row=row, column=1, value=n)
        c.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
        row += 1

    widths = [6, 42, 14, 14, 5, 6, 7, 8, 9, 12, 50, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A6"

    write_carolina_sheet(wb)
    write_twiddy_sheet(wb)
    write_sos_sheet(wb)

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")


def _yes_no(b: bool | None) -> str:
    if b is True:
        return "Yes"
    if b is False:
        return "No"
    return ""


def write_carolina_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Carolina Designs Direct")
    n_cols = len(CD_HEADERS)

    row = 1
    title = ws.cell(row=row, column=1, value="Carolina Designs Realty — 7+ BR, Sat Aug 15 → Sat Aug 22, 2026")
    title.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    row += 1

    sub = ws.cell(row=row, column=1, value="Direct-from-property-manager listings, sorted by recommended → least. Filtered to ≤ $11,000 weekly rate.")
    sub.font = Font(italic=True, color="555555")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    row += 2

    sec = ws.cell(row=row, column=1, value="In-Budget Picks (7 listings of 230 that have Aug 15 availability)")
    sec.font = SECTION_FONT
    sec.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    row += 1

    for i, h in enumerate(CD_HEADERS, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    row += 1

    for entry in CD_PICKS:
        rank, name, town, subdivision, proximity, total, br, fb, hb, sleeps, pets, pool, why, url = entry
        cells = [
            (1, rank, CENTER, None),
            (2, name, WRAP, None),
            (3, town, None, None),
            (4, subdivision, None, None),
            (5, proximity, None, None),
            (6, total, RIGHT, "$#,##0"),
            (7, br, CENTER, None),
            (8, fb, CENTER, None),
            (9, hb, CENTER, None),
            (10, sleeps, CENTER, None),
            (11, _yes_no(pets), CENTER, None),
            (12, _yes_no(pool), CENTER, None),
            (13, why, WRAP, None),
            (14, url, None, None),
        ]
        for col, value, align, fmt in cells:
            c = ws.cell(row=row, column=col, value=value)
            c.border = BORDER
            if align is not None:
                c.alignment = align
            if fmt is not None and value is not None:
                c.number_format = fmt
        url_cell = ws.cell(row=row, column=14)
        url_cell.hyperlink = url
        url_cell.font = LINK_FONT
        url_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        row += 1

    row += 2
    sec2 = ws.cell(row=row, column=1, value="Notes")
    sec2.font = SECTION_FONT
    sec2.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    row += 1
    for n in CD_NOTES:
        c = ws.cell(row=row, column=1, value=n)
        c.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        row += 1

    widths = [6, 28, 18, 24, 26, 14, 5, 7, 7, 7, 8, 10, 60, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


def write_twiddy_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Twiddy Direct")
    n_cols = len(TWIDDY_HEADERS)

    row = 1
    title = ws.cell(row=row, column=1, value="Twiddy & Company — 7+ BR, Sat Aug 15 → Sat Aug 22, 2026")
    title.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    row += 1

    sub = ws.cell(row=row, column=1, value="Direct-from-property-manager listings, ≤ $11,000 weekly rate. Twiddy's 7+ BR inventory skews premium — most exceed budget.")
    sub.font = Font(italic=True, color="555555")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    row += 2

    sec = ws.cell(row=row, column=1, value="In-Budget Picks (1 listing of 368 that have Aug 15 availability and price ≤ $11k)")
    sec.font = SECTION_FONT
    sec.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    row += 1

    for i, h in enumerate(TWIDDY_HEADERS, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    row += 1

    for entry in TWIDDY_PICKS:
        rank, name, town, nbhd, prox, total, br, fb, hb, sleeps, oceanfront, pets, pool, why, url = entry
        cells = [
            (1, rank, CENTER, None),
            (2, name, WRAP, None),
            (3, town, None, None),
            (4, nbhd, None, None),
            (5, prox, None, None),
            (6, total, RIGHT, "$#,##0"),
            (7, br, CENTER, None),
            (8, fb, CENTER, None),
            (9, hb, CENTER, None),
            (10, sleeps, CENTER, None),
            (11, _yes_no(oceanfront), CENTER, None),
            (12, _yes_no(pets), CENTER, None),
            (13, _yes_no(pool), CENTER, None),
            (14, why, WRAP, None),
            (15, url, None, None),
        ]
        for col, value, align, fmt in cells:
            c = ws.cell(row=row, column=col, value=value)
            c.border = BORDER
            if align is not None:
                c.alignment = align
            if fmt is not None and value is not None:
                c.number_format = fmt
        url_cell = ws.cell(row=row, column=15)
        url_cell.hyperlink = url
        url_cell.font = LINK_FONT
        url_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        row += 1

    row += 2
    sec2 = ws.cell(row=row, column=1, value="Notes")
    sec2.font = SECTION_FONT
    sec2.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    row += 1
    for n in TWIDDY_NOTES:
        c = ws.cell(row=row, column=1, value=n)
        c.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        row += 1

    widths = [6, 24, 14, 22, 30, 14, 5, 7, 7, 7, 11, 8, 12, 60, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


def write_sos_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Surf or Sound Direct")
    n_cols = len(SOS_HEADERS)

    row = 1
    title = ws.cell(row=row, column=1, value="Surf or Sound Realty (Hatteras Island) — 7+ BR, Sat Aug 15 → Sat Aug 22, 2026")
    title.font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    row += 1

    sub = ws.cell(row=row, column=1, value="South-OBX coverage: Hatteras Island villages (Rodanthe → Hatteras village). 13 of 13 listings came in under $11k.")
    sub.font = Font(italic=True, color="555555")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    row += 2

    sec = ws.cell(row=row, column=1, value="In-Budget Picks (all 13 listings with Aug 15 availability — every result was under budget)")
    sec.font = SECTION_FONT
    sec.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    row += 1

    for i, h in enumerate(SOS_HEADERS, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    row += 1

    for entry in SOS_PICKS:
        rank, name, town, location, total, br, fb, hb, lid, why = entry
        url = f"{SOS_URL_BASE}/{lid}"
        cells = [
            (1, rank, CENTER, None),
            (2, name, WRAP, None),
            (3, town, None, None),
            (4, location, None, None),
            (5, total, RIGHT, "$#,##0"),
            (6, br, CENTER, None),
            (7, fb, CENTER, None),
            (8, hb, CENTER, None),
            (9, why, WRAP, None),
            (10, url, None, None),
        ]
        for col, value, align, fmt in cells:
            c = ws.cell(row=row, column=col, value=value)
            c.border = BORDER
            if align is not None:
                c.alignment = align
            if fmt is not None and value is not None:
                c.number_format = fmt
        url_cell = ws.cell(row=row, column=10)
        url_cell.hyperlink = url
        url_cell.font = LINK_FONT
        url_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        row += 1

    row += 2
    sec2 = ws.cell(row=row, column=1, value="Notes")
    sec2.font = SECTION_FONT
    sec2.fill = SECTION_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    row += 1
    for n in SOS_NOTES:
        c = ws.cell(row=row, column=1, value=n)
        c.alignment = WRAP
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        row += 1

    widths = [6, 22, 16, 18, 14, 5, 7, 7, 70, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"


if __name__ == "__main__":
    main()
